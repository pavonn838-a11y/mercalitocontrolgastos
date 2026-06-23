import json
import os
import sys
from datetime import datetime, date

sys.path.insert(0, os.path.join(os.getcwd(), "vendor"))

def clean_value(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value

def parse_xls(path):
    import xlrd
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    rows = []
    for r in range(1, sheet.nrows):
        row = {}
        empty = True
        for c, header in enumerate(headers):
            cell = sheet.cell(r, c)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate.xldate_as_datetime(value, book.datemode).date().isoformat()
            elif isinstance(value, str):
                value = value.strip()
            if value not in ("", None):
                empty = False
            row[header or f"Columna {c + 1}"] = clean_value(value)
        if not empty:
            rows.append(row)
    return rows

def parse_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() if v is not None else f"Columna {i + 1}" for i, v in enumerate(rows_iter[0])]
    rows = []
    for values in rows_iter[1:]:
        row = {headers[i]: clean_value(values[i] if i < len(values) else None) for i in range(len(headers))}
        if any(v not in ("", None) for v in row.values()):
            rows.append(row)
    return rows

def main():
    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()
    rows = parse_xls(path) if ext == ".xls" else parse_xlsx(path)
    # Excluir fila resumen tipica del reporte de gastos.
    rows = [r for r in rows if r.get("Sucursal") not in ("", None)]
    print(json.dumps(rows, ensure_ascii=False))

if __name__ == "__main__":
    main()
